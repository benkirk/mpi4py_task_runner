#!/usr/bin/env python3

from mpi4py import MPI
import os
import sys
import pwd, grp
import tempfile
import shutil
import platform
from collections import defaultdict
from typing import NamedTuple
from maxheap import MaxHeap
from datetime import datetime, timezone
from parse_args import parse_options
have_humanfriendly = False
try:
    import humanfriendly
    have_humanfriendly = True
except ImportError:
    pass


################################################################################
def flatten(matrix):
    if matrix: return [item for row in matrix for item in row]
    return None

def format_size(val):
    if have_humanfriendly: return humanfriendly.format_size(val)
    return '{:.5e} bytes'.format(val)

def format_number(val):
    if have_humanfriendly: return humanfriendly.format_number(val)
    return '{:,}'.format(val)

def format_timespan(val):
    if have_humanfriendly: return humanfriendly.format_timespan(val)
    return '{:.1f} seconds'.format(val)



################################################################################
class FileEntry(NamedTuple):
    path   : str
    nbytes : int
    mtime  : float
    ctime  : float
    atime  : float

class DirEntry(NamedTuple):
    path      : str
    nbytes    : int
    nitems    : int
    max_mtime : float
    max_ctime : float
    max_atime : float

class IDCounts:
    def __init__(self):
        self.id = None
        self.name = None
        self.nbytes = 0
        self.nitems = 0
        return

    def set_id(self,id_in):
        if id_in != self.id:
            self.id = id_in
            self._id_to_name()
        return

    def _id_to_name(self):
        assert False
        return

    def to_dict(self):
        d = { 'name'   : self.name,
              #'id'     : self.id,
              'nbytes' : self.nbytes,
              'nitems' : self.nitems }
        return d

class UIDCounts(IDCounts):
    def __init__(self):
        IDCounts.__init__(self)
        return

    def _id_to_name(self):
        try:
            self.name = pwd.getpwuid(self.id).pw_name
        except KeyError:
            self.name = '{}*'.format(self.id)
        return

class GIDCounts(IDCounts):
    def __init__(self):
        IDCounts.__init__(self)
        return

    def _id_to_name(self):
        try:
            self.name = grp.getgrgid(self.id).gr_name
        except KeyError:
            self.name = '{}*'.format(self.id)
        return



################################################################################
class MPIClass:

    tags ={ 'ready'         : 10,
            'execute'       : 11,
            'work_reply'    : 20,
            'work_request'  : 21,
            'work_deny'     : 22,
            'dir_request'   : 30,
            'dir_reply'     : 31,
            'progress'      : 40,
            'terminate'     : 1000 }

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def __init__(self,options=None):
        # initialization, get 'options' data structure from rank 0
        self.comm = MPI.COMM_WORLD
        self.rank   = self.comm.Get_rank()
        self.nranks = self.comm.Get_size()
        self.i_am_root = False if self.rank else True
        self.options = self.comm.bcast(options)
        self.dirs = None
        self.maxnumdirs = 0
        self.num_files = 0
        self.num_dirs = 0
        self.num_items = 0
        self.total_size = 0

        self.st_modes = defaultdict(int)

        self.uids = defaultdict(UIDCounts)
        self.gids = defaultdict(GIDCounts)

        self.top_nitems_dirs   = MaxHeap(self.options.heap_size)
        self.top_nbytes_dirs   = MaxHeap(self.options.heap_size)
        self.top_nbytes_files  = MaxHeap(self.options.heap_size)
        self.oldest_mtime_dirs = MaxHeap(self.options.heap_size)
        self.oldest_atime_dirs = MaxHeap(self.options.heap_size)

        return



    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def __del__(self):
        self.cleanup()
        return



    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def init_local_dirs(self):

        # remember the top 'rundir' where we were launched
        self.rundir = os.getcwd()

        # # get specified local temporary directory, if exists.
        # # SLURM_JOB_TMPFS_TMPDIR, tmpfs ramdisk shared shared by all ranks on node
        # # SLURM_JOB_LOCAL_TMPDIR, /local/.XXXX-user shared by all ranks on node
        # local_topdir = None
        # if not local_topdir: local_topdir = os.getenv('SLURM_JOB_TMPFS_TMPDIR')
        # if not local_topdir: local_topdir = os.getenv('SLURM_JOB_LOCAL_TMPDIR')

        # # local_topdir from slurm is job specific, let's create a subdirectory
        # # for this spefific MPI rank
        # self.local_rankdir = tempfile.mkdtemp(prefix='rank{}_'.format(self.rank),
        #                                       dir=local_topdir)

        return



    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def cleanup(self):
        # # if we set up a local_rankdir, go back to the top workspace 'rundir'
        # # and clean up any temporary leftovers
        # if self.local_rankdir:
        #     os.chdir(self.rundir)
        #     shutil.rmtree(self.local_rankdir,ignore_errors=True)
        #     self.local_rankdir = None
        return



    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def gather_and_sum_dict(self, my_part):
        result = defaultdict(int)
        parts = self.comm.gather(list(my_part.items()))
        if parts:
            for part in parts:
                for k,v in part:
                    result[k] += v

            # sort from largest-val-to-smallest
            # https://stackoverflow.com/questions/613183/how-do-i-sort-a-dictionary-by-value
            result = dict(sorted(result.items(), reverse=True, key=lambda item: item[1]))
        return result



    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def gather_and_sum_ids(self, my_part):
        result = defaultdict(IDCounts)
        parts = self.comm.gather(list(my_part.values()))
        if parts:
            for part in parts:
                for obj in part:
                    k = obj.id
                    result[k].id = obj.id
                    result[k].name = obj.name
                    result[k].nbytes += obj.nbytes
                    result[k].nitems += obj.nitems

            # sort from largest-val-to-smallest
            # https://stackoverflow.com/questions/613183/how-do-i-sort-a-dictionary-by-value
            result = dict(sorted(result.items(), reverse=True, key=lambda item: item[1].nbytes))
        return result



    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def summary(self, verbose=False):

        self.comm.Barrier()

        sep='-'*80

        # gather heaps
        self.top_nitems_dirs.reset   (flatten( self.comm.gather(self.top_nitems_dirs.get_list())))
        self.top_nbytes_dirs.reset   (flatten( self.comm.gather(self.top_nbytes_dirs.get_list())))
        self.top_nbytes_files.reset  (flatten( self.comm.gather(self.top_nbytes_files.get_list())))
        self.oldest_mtime_dirs.reset (flatten( self.comm.gather(self.oldest_mtime_dirs.get_list())))
        self.oldest_atime_dirs.reset (flatten( self.comm.gather(self.oldest_atime_dirs.get_list())))

        self.st_modes   = self.gather_and_sum_dict(self.st_modes)

        self.uids = self.gather_and_sum_ids(self.uids)
        self.gids = self.gather_and_sum_ids(self.gids)

        sys.stdout.flush()

        #--------------------------------------------------
        if verbose:
            for p in range(0,self.nranks):
                self.comm.Barrier()
                sys.stdout.flush()
                if p == self.rank:
                    if self.i_am_root:
                        print(sep)
                    else:
                        print('rank {} / {}, found {:,} files, {:,} dirs'.format(self.rank, platform.node(),
                                                                                 self.num_files, self.num_dirs))
                        for k,v in self.st_modes.items():
                            print('   {:5s} : {:,}'.format(k,v))
                        print('   {:5s} : {}'.format('size',format_size(self.total_size)))


        #------------------------------
        # done with all colletive stuff
        # root rank summarizes results
        if not self.i_am_root: return

        # summarize stat types
        print(('\n'+sep)*3)
        print('Total Count: {} items'.format(format_number(self.progress_counts[0])))
        print('Total Size:  {}'.format(format_size(self.progress_sizes[0])))
        print('Type Counts:')
        for k,v in self.st_modes.items(): print('   {:5s} : {:,}'.format(k, v))

        # summarize by uid/gid
        print('\n' + sep + '\nUsers:\n' + sep)
        for k,v in self.uids.items(): print('{:>12} : {:>10} {:>10}'.format(v.name,format_size(v.nbytes),format_number(v.nitems)))
        print('\n' + sep + '\nGroups:\n' + sep)
        for k,v in self.gids.items(): print('{:>12} : {:>10} {:>10}'.format(v.name,format_size(v.nbytes),format_number(v.nitems)))

        # summarize top files & directories
        print('\n' + sep + '\nLargest Dirs (file count):\n' + sep)
        for idx,de in self.top_nitems_dirs.top(50): print('{:>10} {:>10} {}/'.format(format_number(de.nitems), format_size(de.nbytes), de.path))
        print('\n' + sep + '\nLargest Dirs (size):\n' + sep)
        for idx,de in self.top_nbytes_dirs.top(50): print('{:>10} {:>10} {}/'.format(format_size(de.nbytes), format_number(de.nitems), de.path))
        print('\n' + sep + '\nLargest Files (size):\n' + sep)
        for idx,fe in self.top_nbytes_files.top(50): print('{:>10} {}'.format(format_size(fe.nbytes), fe.path))

        # summarize oldest paths
        print('\n' + sep + '\nOldest Dirs (contents mtimes):\n' + sep)
        for idx,de in self.oldest_mtime_dirs.top(50): print('{} {:>10} {:>10} {}/'.format(datetime.fromtimestamp(de.max_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                                                                                          format_size(de.nbytes), format_number(de.nitems), de.path))
        print('\n' + sep + '\nOldest Dirs (contents atimes):\n' + sep)
        for idx,de in self.oldest_atime_dirs.top(50): print('{}  {:>10} {:>10} {}/'.format(datetime.fromtimestamp(de.max_atime).strftime('%Y-%m-%d %H:%M:%S'),
                                                                                               format_size(de.nbytes), format_number(de.nitems), de.path))

        # write summary file, if requested
        if self.options.summary:
            print('\n--> Writing summary to {}'.format(self.options.summary))
            import pandas as pd

            sheet_prefix = '{} - '.format(self.options.summary_prefix) if self.options.summary_prefix else ''
            mode = 'a' if os.path.exists(self.options.summary) else 'w'
            if_sheet_exists = None if 'w' == mode else 'replace'

            with pd.ExcelWriter(self.options.summary,
                                mode=mode, if_sheet_exists=if_sheet_exists) as writer:
                # directories
                l = []
                for idx,de in self.top_nitems_dirs.top(self.options.heap_size): l.append(de)
                for idx,de in self.top_nbytes_dirs.top(self.options.heap_size): l.append(de)
                #for idx,de in self.oldest_mtime_dirs.top(self.options.heap_size): l.append(de)
                #for idx,de in self.oldest_ctime_dirs.top(self.options.heap_size): l.append(de)
                #for idx,de in self.oldest_atime_dirs.top(self.options.heap_size): l.append(de)
                l = sorted(set(l), key = lambda x: x.nitems, reverse=True)
                df = pd.DataFrame(l)
                for ts in ['max_mtime', 'max_ctime', 'max_atime']: df[ts] = pd.to_datetime(df[ts], unit='s')
                df['size'] = df['nbytes'].apply(lambda x: format_size(x))
                df = df[['path','nitems','size','nbytes','max_mtime','max_atime']] # <-- rearrange, drop ctime (don't report what will likely be confusing information)
                #print(df)
                df.to_excel(writer, sheet_name=sheet_prefix+'Directories', index=False)
                del l, df

                # files
                l = []
                for idx,fe in self.top_nbytes_files.top(self.options.heap_size): l.append(fe)
                l = sorted(set(l), key = lambda x: x.nbytes, reverse=True)
                df = pd.DataFrame(l)
                for ts in ['mtime', 'ctime', 'atime']: df[ts] = pd.to_datetime(df[ts], unit='s')
                df['size'] = df['nbytes'].apply(lambda x: format_size(x))
                df = df[['path','size','nbytes','mtime','atime']] # <-- rearrange, drop ctime (don't report what will likely be confusing information)
                #print(df)
                df.to_excel(writer, sheet_name=sheet_prefix+'Files', index=False)
                del l, df

                # UID counts
                udf = pd.DataFrame.from_records([id.to_dict() for id in self.uids.values()])
                udf['groupname'] = udf['name'].apply(lambda x: None)
                udf.rename(columns={'name': 'username'},inplace=True)
                udf['size'] = udf['nbytes'].apply(lambda x: format_size(x))
                udf = udf[['username', 'groupname', 'size', 'nbytes', 'nitems']]

                # GID counts
                gdf = pd.DataFrame.from_records([id.to_dict() for id in self.gids.values()])
                gdf['username'] = gdf['name'].apply(lambda x: None)
                gdf.rename(columns={'name': 'groupname'},inplace=True)
                gdf['size'] = gdf['nbytes'].apply(lambda x: format_size(x))
                gdf = gdf[['username', 'groupname', 'size', 'nbytes', 'nitems']]

                # combined sheet
                df = pd.concat([udf,gdf],ignore_index=True)
                #print(df)
                df.to_excel(writer, sheet_name=sheet_prefix+'Users & Groups', index=False)
                del udf, gdf, df


            # Ok, open the file and set some formatting
            from openpyxl import Workbook, load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Alignment, Font

            wb = load_workbook(self.options.summary)
            #print(wb.sheetnames)

            for ws in wb.worksheets:
                #print(ws)
                ws.freeze_panes = 'A2'
                # Iterate over all columns and adjust their widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    #print(column[0].value)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)

                            cell.font = Font(name='Courier')

                            if 'size' in column[0].value:
                                cell.alignment = Alignment(horizontal='right')
                            if 'nitems' in column[0].value:
                                cell.number_format = '#,###'
                            if 'time' in column[0].value:
                                cell.number_format = 'yyyy-mm-dd h:mm:ss'

                        except:
                            pass

                    adjusted_width = (max_length + 2) * 1.2

                    ws.column_dimensions[column_letter].width = adjusted_width

                    # special widths
                    if 'time' in column[0].value:
                        ws.column_dimensions[column_letter].width = 20
                    elif 'path' in column[0].value:
                        ws.column_dimensions[column_letter].width = min(adjusted_width,140)
                    elif 'nbytes' in column[0].value:
                        ws.column_dimensions[column_letter].hidden = True

                    # column label formatting
                    column[0].font = Font(bold=True, size=14)
                    column[0].alignment = Alignment(horizontal='center')
                    #column[0].value = column[0].value.replace('_',' ').title()

            # Save the workbook
            wb.save(self.options.summary)
            print('--> Done at {}'.format(datetime.now().isoformat(sep=' ', timespec='seconds')))

        return
